import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook

# Função para atualizar o arquivo
def atualizar_arquivo():
    try:
        # Lê o arquivo de entrada e extrai as duas colunas desejadas
        df = pd.read_excel(file_input.get(), usecols='H, G', skiprows=1)

        # Aplica o aumento percentual sobre a coluna "Custo"
        porcentagem = float(porcentagem_entry.get().replace(",", ".")) # Converte a porcentagem para float
        df["Custo"] = df["Custo"] * (1 + porcentagem/100)

        # Abre o arquivo de destino em modo de leitura
        book = load_workbook(file_output.get())

        # Seleciona a planilha "Anúncios" para atualização
        sheet = book["Anúncios"]

        # Atualiza as células E3:EX com os valores da coluna "Estoque" do arquivo de entrada
        for i, estoque in enumerate(df["Estoque"], start=4):
            sheet.cell(row=i, column=5, value=estoque)

        # Atualiza as células F3:FX com os valores da coluna "Custo" do arquivo de entrada
        for i, custo in enumerate(df["Custo"], start=4):
            sheet.cell(row=i, column=6, value=custo)

        # Salva as alterações no arquivo de destino
        book.save(file_output.get())

        # Mostra uma mensagem de sucesso
        messagebox.showinfo("Sucesso", "Arquivo atualizado com sucesso!")

    except Exception as e:
        # Mostra uma mensagem de erro
        messagebox.showerror("Erro", str(e))

    finally:
        # Fecha a janela principal
        root.destroy()

# Cria uma janela
root = tk.Tk()
root.title("Atualização de Arquivo")

# Define o estilo da janela
root.configure(bg="#F5F5F5")

# Define o tamanho e posição da janela
largura_janela = 650
altura_janela = 300
largura_tela = root.winfo_screenwidth()
altura_tela = root.winfo_screenheight()
posx = largura_tela/2 - largura_janela/2
posy = altura_tela/2 - altura_janela/2
root.geometry("%dx%d+%d+%d" % (largura_janela, altura_janela, posx, posy))

# Cria os widgets para seleção dos arquivos
tk.Label(root, text="Arquivo de Entrada:", font=("Arial", 12)).grid(row=0, column=0, sticky="w", padx=20, pady=10)
file_input_frame = tk.Frame(root)
file_input_frame.grid(row=0, column=1, pady=10)

file_input = tk.Entry(file_input_frame, width=30, font=("Arial", 12))
file_input.grid(row=0, column=0)

browse_input_button = tk.Button(file_input_frame, text="Procurar", font=("Arial", 12), command=lambda: file_input.insert(
    0, filedialog.askopenfilename()))
browse_input_button.grid(row=0, column=1, padx=10)

tk.Label(root, text="Arquivo de Saída:", font=("Arial", 12)).grid(row=1, column=0, sticky="w", padx=20, pady=10)
file_output_frame = tk.Frame(root)
file_output_frame.grid(row=1, column=1, pady=10)

file_output = tk.Entry(file_output_frame, width=30, font=("Arial", 12))
file_output.grid(row=0, column=0)

browse_output_button = tk.Button(file_output_frame, text="Procurar", font=("Arial", 12), command=lambda: file_output.insert(
    0, filedialog.asksaveasfilename(defaultextension=".xlsx")))
browse_output_button.grid(row=0, column=1, padx=10)

# Cria um widget para escolher a porcentagem de aumento
tk.Label(root, text="Porcentagem de Aumento:", font=("Arial", 12)).grid(row=2, column=0, sticky="w", padx=20, pady=10)
porcentagem_frame = tk.Frame(root)
porcentagem_frame.grid(row=2, column=1)

porcentagem_entry = tk.Entry(porcentagem_frame, width=10, font=("Arial", 12))
porcentagem_entry.grid(row=0, column=0)

porcentagem_label = tk.Label(porcentagem_frame, text="%", font=("Arial", 12))
porcentagem_label.grid(row=0, column=1)

# Cria um widget vazio para adicionar espaço entre o campo de porcentagem e o botão de atualizar
tk.Label(root, text="", font=("Arial", 12)).grid(row=3, column=1)

# Cria um botão para atualizar o arquivo
update_button = tk.Button(root, text="Atualizar", font=("Arial", 12), command=atualizar_arquivo)
update_button.grid(row=4, column=1, pady=20)

# Inicia o loop principal da interface gráfica
root.mainloop()